"""
	*.txt文件----->*.xlsx文件
"""
"""
	读取选项比例文件,返回{quest:{"ans1":ans1perc,"ans2":ans02perc}}字典
		getpercentage(filepath,peopleNum)
	选项分类,获取积极选项
		anaiyans(gqdict)
	写入输出excel文件
		outputexcel(filepath)
openpyxl读取格式:
		column01  column02  column03  column04  column05
	row01
	row02
	row03
	row04
	row05
	row06

问题格式:
	01.xxxxx(问题)
		> yyy(答案)
		> zzz(--->perc:0.3)(答案,百分比)
	02.xxxx
		> uuu
		> vvv(--->perc:0.2)
		> www(--->perc:0.4)
		...
	...
"""
import openpyxl
import os
import re

# 读取选项比例文件,peopleNum:总人数
def getpercentage(filepath,peopleNum):
	returndict = dict()
	with open(filepath,"r",encoding="utf-8") as gpfile:	# getpercentagefile
		gpList = gpfile.readlines()
	nowquest = str()	# 当前的问题
	# 获取已赋值的选项
	for line in gpList:
		matchquest = re.match(r"^(\t*)(\d+)(.*)$",line)	# 匹配问题
		matchans = re.match(r"^(\t*)(>)(.*)$",line) # 匹配选项
		matchperc = re.match(r"^(\t*)(>)(.*)\(--->perc:(.*)\)$",line) # 匹配选项比例
		if matchquest:	# 匹配到问题
			quest = str(matchquest.group(2)+matchquest.group(3))
			nowquest = quest
			returndict[quest] = dict()
		if matchans:
			ans = re.sub(r"\(--->perc:(.*)\)","",matchans.group(3))
			if matchperc:	# 匹配到比例
				percent = int(float(matchperc.group(4))*peopleNum)
			else:
				percent = int(0)
			returndict[nowquest][ans] = percent
	# 修正总人数
	questkeys = list(returndict.keys())
	for quest in questkeys:
		anskeys = list(returndict[quest].keys())	# 选项列表
		ansdict = returndict[quest]				# 选项人数
		emptyans = str()							# 人数为0的选项
		totalnum = int()							# "人数不为0的选项"的总人数
		for ans in anskeys:
			totalnum = totalnum+ansdict[ans]
			if ansdict[ans]==0:
				emptyans = ans
		if totalnum < peopleNum:					# 为单选题
			returndict[quest][emptyans] = peopleNum - totalnum
		else:										# 为多选不用管
			pass							
	return returndict

# 写入输出excel文件,startline:开始行,startrow:开始列,ansSplitFlag:多选项问题答案的分割符默认为";"
def outputexcel(filepath,sheetName,questdict,peopleNum,ansSplitFlag=";",startcolumn=1,startrow=1):
	"""
		1行(1line):写入问题
		2行-(总人数+2)行:写入答案
		按列进行读写数据
	"""
	workbook = openpyxl.Workbook()							# 创建新的excel文件
	worksheet = workbook.create_sheet(sheetName)			# 创建新的表格
	workbook.save(filepath)									# 保存文件
	workbook = openpyxl.load_workbook(filename = filepath)	# 读取文件
	worksheet = workbook[sheetName]							# 读取表格
	questList = list(questdict.keys())						# 问题列表
	maxrow = peopleNum + startrow							# 允许写的最大行数,根据人数而定,从第二行写数据
	maxcolumn = len(questList)								# 允许写的最大列数,根据问题的数目而定
	columnindex = startcolumn								# 初始列标记
	rowindex = startrow										# 初始行标记
	# 读取列(line)--->读取问题的人数--->读取行(row)写入数据
	for column in range(maxcolumn):							
		quest = questList[column]						# rowindex对应的问题
		ansdict = questdict[quest]						# 答案及其对应人数的列表
		anslist = list(ansdict.keys())					# 答案所构成的列表
		rowNum = len(anslist)							# 答案列的行数
		worksheet.cell(column=columnindex,row=startrow).value = quest 		# 写入问题
		ansrowindex = startrow + 1						# 问题的初始行
		for ans in anslist:								# 遍历问题,ans:问题
			ansMan = ansdict[ans]						# 问题的人数
			for row in range(ansMan):					# 遍历列(row)
				cellread = worksheet.cell(column=columnindex,row=ansrowindex).value	# 读取(rowindex,lineindex)处的数据
				if cellread and (ans not in cellread):
					writeans = cellread + ansSplitFlag + ans  								# 为多选的选项
				else:
					writeans = ans		
				worksheet.cell(column=columnindex,row=ansrowindex).value = writeans  	# 写入答案
				ansrowindex += 1
				if ansrowindex > maxrow:					# 超过最大maxrow则重置rowindex
					ansrowindex = startrow + 1				# 重置rowindex至问题下一行
		columnindex += 1
	# 遍历数据
	for col in range(maxcolumn):
		for row in range(maxrow):
			print("column列:{0}-row行:{1}----->data:{2}".format(col+1,row+1,worksheet.cell(row=row+1,column=col+1).value))
	#关闭工作文件
	workbook.save(filepath)

# 主控制函数,questfile:问题答案文本文件 excelfile:输出excel文件 sheetname:写入数据的表格名 peopleNum:回答问题的总人数
def run(questfile,excelfile,sheetname,peopleNum,ansSplitFlag=";",startcolumn=1,startrow=1):
	getpercdict = getpercentage(filepath=questfile,peopleNum=peopleNum)
	outputexcel(filepath=excelfile,sheetName=sheetname,questdict=getpercdict,peopleNum=peopleNum,ansSplitFlag=ansSplitFlag,startcolumn=startcolumn,startrow=startrow)

if __name__ == "__main__":
	run("2021.09.02问卷问题-perctage.txt","./2021.09.02问卷问题-perctage.xlsx","mainsheet",120)